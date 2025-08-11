import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import PatternFill
from datetime import datetime

# Page config
st.set_page_config(page_title="SSV Automation Tool", layout="wide", page_icon="📊")

# Enhanced Header with Gradient and Shadow
st.markdown("""
    <style>
    .main {background-color: #f9f9f9;}
    .stApp {font-family: 'Segoe UI';}
    .title {
        font-size: 40px;
        font-weight: 800;
        background: linear-gradient(to right, #ff6a00, #ee0979);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        padding: 10px;
        text-shadow: 1px 1px 2px #ccc;
    }
    .subtitle {
        font-size: 22px;
        color: #555;
        text-align: center;
        margin-bottom: 20px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="title">📊 SSV Factors Automation Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Generate Sheet3 with grouped policy values and identifiers</div>', unsafe_allow_html=True)

with st.expander("ℹ️ Instructions", expanded=False):
    st.markdown("""
    - Upload your Excel file containing the `SSV_Factors` sheet.
    - Enter product details and configuration.
    - Click generate to download the processed Excel file containing only Sheet3.
    """)

# Input section
with st.container():
    col1, col2 = st.columns(2)
    with col1:
        uploaded_file = st.file_uploader("📁 Upload Excel File", type=["xlsx"])
        product = st.text_input("🏷️ Product Name - T5687", placeholder="e.g., T36A")
        start_indicator = st.text_input("🔢 Start Indicator -TZA78", placeholder="e.g., 3")
    with col2:
        insp_start = st.number_input("📍 INSPRM Start Column - TZA66", min_value=1, max_value=99, value=8)
        insp_end = st.number_input("📍 INSPRM End Column -TZA66", min_value=1, max_value=99, value=55)
        multiplier = st.number_input("✖️ Multiplication Factor", min_value=1.0, value=10000.0)

# Process and generate
if uploaded_file and product and start_indicator:
    if len(start_indicator) == 1:
        start_indicator = f"0{start_indicator}"

    with st.spinner("🔄 Processing your file..."):
        ssv_factors = pd.read_excel(uploaded_file, sheet_name="SSV_Factors", engine="openpyxl")
        policy_terms = ssv_factors.columns[1:].tolist()
        policy_durations = ssv_factors.iloc[:, 0].tolist()
        values_matrix = ssv_factors.iloc[:, 1:].values

        # Sheet1 (processed internally)
        sheet1_rows = []
        for j, term in enumerate(policy_terms):
            for i, duration in enumerate(policy_durations):
                indicator = int(start_indicator) + i
                indicator_str = f"{indicator:02d}"
                value = values_matrix[i][j]
                identifier = f"{product}{int(term)}{indicator_str}"
                row = [indicator_str, duration, product, value, int(term), identifier, identifier]
                sheet1_rows.append(row)

        sheet1_df = pd.DataFrame(sheet1_rows, columns=["Indicator", "Policy duration (in months)", "Product", "Value", "Policy Term", "Identifier", "Identifier2"])

        # Sheet2 (processed internally)
        sheet2_df = sheet1_df.copy()
        sheet2_df["Policy Value"] = sheet2_df["Value"].fillna(0) * multiplier
        sheet2_df = sheet2_df[["Policy Value", "Identifier"]]

        # Sheet3
        sheet3_columns = ["ITEMITEM"] + ["INSTPR", "MFACTHM", "MFACTHY", "MFACTM", "MFACTQ", "MFACTW", "MFACT2W", "MFACT4W", "MFACTY", "PREM_UNIT", "UNIT", "DISCCNTMETH"]
        sheet3_columns += [f"INSPRM{i:02d}" for i in range(1, 100)]
        sheet3_columns += ["USER_PROFILE", "JOB_NAME", "DATIME", "INSPREM"]
                                                    
        current_datetime = datetime.now().strftime("%Y-%m-%d-%H.%M.%S.000000")

        sheet3_rows = []
        for _, row in sheet2_df.iterrows():
            identifier = row["Identifier"]
            value = row["Policy Value"]
            row_data = [identifier] + [0]*9 + [100, 0, ""]
            for i in range(1, 100):
                if insp_start <= i <= insp_end:
                    row_data.append(value)
                else:
                    row_data.append(0)
            row_data += ["LAAKRUTI", "L2UPLDPRM", current_datetime, 0]
            sheet3_rows.append(row_data)

        sheet3_df = pd.DataFrame(sheet3_rows, columns=sheet3_columns)

        # Replace blanks only in INSPRM columns (NaN or empty string)
        insp_columns = [col for col in sheet3_df.columns if col.startswith("INSPRM")]
        for col in insp_columns:
            sheet3_df[col] = sheet3_df[col].apply(lambda x: 0 if pd.isna(x) or x == '' else x)

        # Save to Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet3_df.to_excel(writer, sheet_name="Sheet3", index=False)

            wb = writer.book
            ws = wb["Sheet3"]
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")

    st.success("🎉 Sheet generated successfully!")
    st.download_button("📥 Download Excel File", data=output.getvalue(), file_name="generated_sheet.xlsx")

else:
    st.warning("⚠️ Please upload a file and fill all inputs to proceed.")


